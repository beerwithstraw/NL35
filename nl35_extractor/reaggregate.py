"""
Re-aggregation Script for NL-35.
Rebuilds Master_Data sheet from individual company verification sheets.
"""

import sys
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import click
import openpyxl
from rich.console import Console

sys.path.insert(0, str(Path(__file__).parent))

from extractor.models import NL35Extract, NL35Data
from config.row_registry import NL35_LOB_ORDER, NL35_LOB_DISPLAY_NAMES
from config.settings import PERIOD_METRIC_KEYS
from validation.checks import run_validations, write_validation_report, build_validation_summary_table
from output.excel_writer import save_workbook

console = Console()
logger = logging.getLogger(__name__)


def _parse_sheet_to_extract(ws) -> Optional[NL35Extract]:
    """Parse a verification sheet back into NL35Extract."""
    try:
        meta_cell = ws.cell(row=2, column=1).value
        if not meta_cell or "Quarter:" not in meta_cell:
            return None
        parts = {p.split(":")[0].strip(): p.split(":")[1].strip() for p in meta_cell.split("|")}
        quarter = parts.get("Quarter", "Qx")
        year = parts.get("Year", "xxxxxx")
        source_file = parts.get("Source", "Unknown.pdf")

        title = ws.cell(row=1, column=1).value
        company_name = title.replace("VERIFICATION:", "").strip() if title else "Unknown"

        from config.company_registry import COMPANY_DISPLAY_NAMES
        company_key = "unknown"
        for k, v in COMPANY_DISPLAY_NAMES.items():
            if v == company_name:
                company_key = k
                break

        extract = NL35Extract(
            source_file=source_file,
            company_key=company_key,
            company_name=company_name,
            form_type="NL35",
            quarter=quarter,
            year=year,
        )

        # Read data grid (header at row 4, data from row 5)
        header_row = 4
        for ri, lob in enumerate(NL35_LOB_ORDER, header_row + 1):
            lob_vals = {}
            for ci, key in enumerate(PERIOD_METRIC_KEYS, 2):
                val = ws.cell(row=ri, column=ci).value
                try:
                    lob_vals[key] = float(val) if val is not None else None
                except (ValueError, TypeError):
                    lob_vals[key] = None
            extract.data.data[lob] = lob_vals

        return extract
    except Exception as e:
        logger.error(f"Failed to parse sheet {ws.title}: {e}")
        return None


@click.command()
@click.option("--workbook", "-w", required=True, help="Path to the Excel master workbook", type=click.Path(exists=True))
@click.option("--backup/--no-backup", default=True)
def reaggregate(workbook, backup):
    """Rebuild Master_Data sheet from individual company sheets."""
    console.print(f"[bold blue]Re-aggregating workbook:[/bold blue] {workbook}")

    if backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = Path(workbook).with_name(f"{Path(workbook).stem}_backup_{timestamp}.xlsx")
        shutil.copy2(workbook, backup_path)
        console.print(f"[dim]Backup created: {backup_path.name}[/dim]")

    try:
        wb = openpyxl.load_workbook(workbook)
        extractions = []

        for sheet_name in wb.sheetnames:
            if sheet_name in ["Master_Data", "_meta", "Validation_Summary", "Validation_Detail"]:
                continue
            ws = wb[sheet_name]
            extract = _parse_sheet_to_extract(ws)
            if extract:
                extractions.append(extract)
                console.print(f"  [green]\u2713[/green] Parsed {sheet_name}")
            else:
                console.print(f"  [red]\u2717[/red] Skipped {sheet_name}")

        if not extractions:
            console.print("[bold red]No valid company sheets found.[/bold red]")
            return

        val_results = run_validations(extractions)
        report_path = Path(workbook).parent / "validation_report_nl35_reagg.csv"
        write_validation_report(val_results, str(report_path))

        save_workbook(extractions, workbook, stats={"files_processed": len(extractions), "files_succeeded": len(extractions)})

        console.print(f"\n[bold green]Re-aggregation successful![/bold green]")
        console.print(build_validation_summary_table(val_results))

    except Exception as e:
        console.print(f"[bold red]Critical error:[/bold red] {e}")
        sys.exit(1)


if __name__ == "__main__":
    reaggregate()
