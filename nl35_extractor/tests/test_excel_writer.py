"""
test_excel_writer.py — NL-35 Excel output tests.
"""

import pytest
from extractor.models import NL35Extract, NL35Data
from output.excel_writer import save_workbook, _write_verification_sheet
from openpyxl import Workbook


def _make_extract():
    ext = NL35Extract(
        source_file="test.pdf",
        company_key="bajaj_allianz",
        company_name="Bajaj Allianz General Insurance Company Limited",
        form_type="NL35",
        quarter="Q3",
        year="202526",
    )
    ext.data.data["fire"] = {
        "cy_qtr_premium": 53116.14,
        "cy_qtr_policies": 703285.0,
        "py_qtr_premium": 50053.40,
        "py_qtr_policies": 629160.0,
        "cy_ytd_premium": 227419.19,
        "cy_ytd_policies": 2030317.0,
        "py_ytd_premium": 199355.36,
        "py_ytd_policies": 1961031.0,
    }
    return ext


def test_save_workbook_creates_file(tmp_path):
    output = str(tmp_path / "test_output.xlsx")
    ext = _make_extract()
    save_workbook([ext], output)
    import os
    assert os.path.exists(output)


def test_save_workbook_has_master_data_sheet(tmp_path):
    from openpyxl import load_workbook
    output = str(tmp_path / "test_output.xlsx")
    ext = _make_extract()
    save_workbook([ext], output)
    wb = load_workbook(output)
    assert "Master_Data" in wb.sheetnames


def test_save_workbook_master_data_has_headers(tmp_path):
    from openpyxl import load_workbook
    from config.settings import MASTER_COLUMNS
    output = str(tmp_path / "test_output.xlsx")
    ext = _make_extract()
    save_workbook([ext], output)
    wb = load_workbook(output)
    ws = wb["Master_Data"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(MASTER_COLUMNS) + 1)]
    assert headers == MASTER_COLUMNS


def test_verification_sheet_written(tmp_path):
    from openpyxl import load_workbook
    output = str(tmp_path / "test_output.xlsx")
    ext = _make_extract()
    save_workbook([ext], output)
    wb = load_workbook(output)
    # Should have a verification sheet for Bajaj
    company_sheets = [s for s in wb.sheetnames if "Bajaj" in s or "bajaj" in s.lower()]
    assert len(company_sheets) >= 1


def test_write_verification_sheet_has_lob_names():
    wb = Workbook()
    ws = wb.active
    ext = _make_extract()
    _write_verification_sheet(ws, ext)
    # Row 5 col 1 should be "Fire"
    assert ws.cell(row=5, column=1).value == "Fire"


def test_write_verification_sheet_fire_cy_qtr_premium():
    wb = Workbook()
    ws = wb.active
    ext = _make_extract()
    _write_verification_sheet(ws, ext)
    # Row 5, col 2 = CY_Qtr_Premium for fire
    val = ws.cell(row=5, column=2).value
    assert val == pytest.approx(53116.14)
