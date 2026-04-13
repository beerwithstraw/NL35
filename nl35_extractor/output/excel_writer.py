"""
Excel Writer for NL-35 Quarterly Business Returns.

Output shape: one row per company per LOB (15 rows per company).
Each row has 8 data columns (4 periods × 2 metrics).
Policy counts are formatted as integers; premiums as 2-decimal floats.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    MASTER_COLUMNS,
    PERIOD_METRIC_KEYS,
    PERIOD_ROW_MAP,
    EXTRACTOR_VERSION,
    NUMBER_FORMAT,
    INTEGER_FORMAT,
    LOW_CONFIDENCE_FILL_COLOR,
    company_key_to_pascal,
)
from config.row_registry import NL35_LOB_ORDER, NL35_LOB_DISPLAY_NAMES
from config.company_metadata import get_metadata
from config.lob_metadata import get_lob_particulars, get_grouped_lob
from extractor.models import NL35Extract

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_CENTER_ALIGN_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_META_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_POLICY_COL = "No_of_Policies"
_PREMIUM_COL = "Premium"


def _year_code_to_fy_end(year_code: str) -> str:
    s = str(year_code).strip()
    if len(s) == 8:
        return s[4:]
    if len(s) == 6:
        return f"20{s[4:]}"
    return s


def _write_master_data(ws, extractions: List[NL35Extract], existing_rows: Optional[List[list]] = None):
    # Header row
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    # Preserved rows from previous runs
    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                col_name = MASTER_COLUMNS[col_idx - 1]
                if col_name == _POLICY_COL:
                    cell.number_format = INTEGER_FORMAT
                elif col_name == _PREMIUM_COL:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    # New extractions — each LOB × 4 period combinations = 4 rows
    for extract in extractions:
        meta = get_metadata(extract.company_key)
        fy_end = _year_code_to_fy_end(extract.year)

        for lob in NL35_LOB_ORDER:
            lob_vals = extract.data.data.get(lob)
            if lob_vals is None:
                continue

            for year_info, quarter_info, prem_key, pol_key in PERIOD_ROW_MAP:
                row_data = {
                    "LOB_PARTICULARS":      get_lob_particulars(lob),
                    "Grouped_LOB":          get_grouped_lob(lob),
                    "Company_Name":         meta["company_name"],
                    "Company":              meta["sorted_company"],
                    "NL":                   extract.form_type,
                    "Quarter":              extract.quarter,
                    "Year":                 fy_end,
                    "Year_Info":            year_info,
                    "Quarter_Info":         quarter_info,
                    "Sector":               meta["sector"],
                    "Industry_Competitors": meta["competitors"],
                    "GI_Companies":         "GI Company",
                    "No_of_Policies":       lob_vals.get(pol_key),
                    "Premium":              lob_vals.get(prem_key),
                    "Source_File":          extract.source_file,
                }

                for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=row_data[col_name])
                    if col_name == _POLICY_COL:
                        cell.number_format = INTEGER_FORMAT
                    elif col_name == _PREMIUM_COL:
                        cell.number_format = NUMBER_FORMAT

                current_row += 1


def _write_verification_sheet(ws, extract: NL35Extract):
    """Per-company verification grid: LOBs as rows, period-metrics as columns."""
    total_cols = 9  # 1 label + 8 data

    # Row 1: company name title
    cell = ws.cell(row=1, column=1, value=extract.company_name)
    cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Row 2: subtitle
    subtitle = f"NL-35  |  Quarter: {extract.quarter}  |  Year: {extract.year}  |  Source: {extract.source_file}"
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = Font(italic=True, size=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    # Row 4: header
    header_row = 4
    cell = ws.cell(row=header_row, column=1, value="Line of Business")
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = _CENTER_ALIGN_WRAP

    col_headers = [
        "CY Qtr\nPremium (₹L)", "CY Qtr\nPolicies",
        "PY Qtr\nPremium (₹L)", "PY Qtr\nPolicies",
        "CY YTD\nPremium (₹L)", "CY YTD\nPolicies",
        "PY YTD\nPremium (₹L)", "PY YTD\nPolicies",
    ]
    for ci, hdr in enumerate(col_headers, 2):
        cell = ws.cell(row=header_row, column=ci, value=hdr)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN_WRAP

    ws.freeze_panes = "A5"

    # Data rows
    for ri, lob in enumerate(NL35_LOB_ORDER, header_row + 1):
        ws.cell(row=ri, column=1, value=NL35_LOB_DISPLAY_NAMES.get(lob, lob))
        lob_vals = extract.data.data.get(lob, {})
        for ci, key in enumerate(PERIOD_METRIC_KEYS, 2):
            val = lob_vals.get(key)
            cell = ws.cell(row=ri, column=ci, value=val)
            if val is not None:
                if "policies" in key:
                    cell.number_format = INTEGER_FORMAT
                else:
                    cell.number_format = NUMBER_FORMAT

    # Column widths
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _write_meta_sheet(ws, extractions: List[NL35Extract], stats: Dict[str, Any]):
    companies = sorted(set(e.company_name for e in extractions))
    quarters = sorted(set(f"{e.quarter}_{e.year}" for e in extractions))
    data = [
        ["Key", "Value"],
        ["extraction_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version", EXTRACTOR_VERSION],
        ["files_processed", stats.get("files_processed", 0)],
        ["files_succeeded", stats.get("files_succeeded", 0)],
        ["files_failed", stats.get("files_failed", 0)],
        ["companies", ", ".join(companies)],
        ["quarters", ", ".join(quarters)],
    ]
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL


def _sheet_name_for(extract: NL35Extract) -> str:
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


def save_workbook(extractions: List[NL35Extract], output_path: str, stats: Optional[Dict[str, Any]] = None):
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []

    if output_file.exists():
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)

        new_files = {e.source_file for e in extractions}

        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]

            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                try:
                    sf_idx = headers.index("Source_File")
                except ValueError:
                    sf_idx = None

                if sf_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if row[sf_idx] is None:
                            continue
                        if row[sf_idx] not in new_files:
                            existing_rows.append(list(row))
            else:
                logger.warning("Existing Master_Data has different column layout — discarding old rows.")

            del wb["Master_Data"]

        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]
        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        wb = Workbook()
        wb.remove(wb.active)

    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows)

    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")


def write_validation_summary_sheet(report_path: str, master_path: str):
    import pandas as pd
    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status",
        aggfunc="size",
        fill_value=0,
    ).reset_index()
    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0
    summary["Files_Processed"] = 1
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    cols = ["Company", "Quarter", "Year", "Files_Processed", "PASS", "SKIP", "WARN", "FAIL"]
    summary["Total_Checks"] = summary[["PASS", "SKIP", "WARN", "FAIL"]].sum(axis=1)
    cols.insert(4, "Total_Checks")
    summary = summary[cols]
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, master_path: str):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)
    cols_map = {
        "company": "Company", "quarter": "Quarter", "year": "Year",
        "lob": "LOB", "check_name": "Check_Name",
        "status": "Status", "expected": "Expected", "actual": "Actual",
        "delta": "Delta", "note": "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values(by="Status").reset_index(drop=True)

    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
        detail.to_excel(writer, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(master_path)
    ws = wb["Validation_Detail"]
    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col = list(cols_map.values()).index("Status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if status_val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")
